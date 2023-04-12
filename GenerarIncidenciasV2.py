import mysql.connector
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

def archiNoExiste(mycursor):
    libro = openpyxl.Workbook()
    # Selecciona la hoja de Excel
    hoja_excel = libro.active
    # Crea encabezados para las columnas
    hoja_excel.cell(row=1, column=1, value="Fecha").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=2, value="Hora Entrada").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=3, value="Hora Salida").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=4, value="Alias").alignment = Alignment(horizontal="center")
    # Agrega los datos a la hoja de Excel
    fila = 2
    for x in mycursor:
        if len(x[0]) == 8:
            fecha = datetime.datetime.strptime(x[0], '%Y%m%d').strftime('%Y/%m/%d')
            hora_entrada = datetime.datetime.strptime(x[1], '%H%M%S').strftime('%H:%M:%S')
            hora_salida = datetime.datetime.strptime(x[2], '%H%M%S').strftime('%H:%M:%S')
            hoja_excel.cell(row=fila, column=1, value=fecha)
            hoja_excel.cell(row=fila, column=2, value=hora_entrada)
            hoja_excel.cell(row=fila, column=3, value=hora_salida)
            hoja_excel.cell(row=fila, column=4, value=x[3])
            fila += 1
        else:
            print(f"Error: Fecha inválida ({x[0]})")

    # Ajusta el ancho de las columnas
    for col in hoja_excel.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja_excel.column_dimensions[column].width = adjusted_width

    # Guarda el archivo de Excel
    libro.save("C:\\AOBREGON\\" + nombre_archivo)

    return 

def archiExiste(mycursor):
    libro = openpyxl.load_workbook("C:\\AOBREGON\\"+nombre_archivo)
    hoja_excel = libro.active
    fila = 1 + hoja_excel.max_row
    for x in mycursor:
        if len(x[0]) == 8:
            fecha = datetime.datetime.strptime(x[0], '%Y%m%d').strftime('%Y/%m/%d')
            hora_entrada = datetime.datetime.strptime(x[1], '%H%M%S').strftime('%H:%M:%S')
            hora_salida = datetime.datetime.strptime(x[2], '%H%M%S').strftime('%H:%M:%S')
            hoja_excel.cell(row=fila, column=1, value=fecha)
            hoja_excel.cell(row=fila, column=2, value=hora_entrada)
            hoja_excel.cell(row=fila, column=3, value=hora_salida)
            hoja_excel.cell(row=fila, column=4, value=x[3])
            fila += 1
        else:
            print(f"Error: Fecha inválida ({x[0]})")

    # Ajusta el ancho de las columnas
    for col in hoja_excel.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja_excel.column_dimensions[column].width = adjusted_width

    # Guarda el archivo de Excel
    libro.save("C:\\AOBREGON\\" + nombre_archivo)



    print("La hoja tiene", fila, "filas con datos.")
    return

# Configura la conexión
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="contraseña",
    database="base de datos"
)

# Obtiene el número de la semana actual
semana_actual = datetime.date.today().isocalendar()[1]

# Nombre del archivo de Excel
nombre_archivo = f"ErroresFichajeSemana{semana_actual}.xlsx"

# Consulta para obtener las incidencias
consulta = "SELECT Fecha, MIN(Hora) as HoraEntrada, MAX(Hora) as HoraSalida, Alias FROM newschema.fichajes1 JOIN newschema.empleados1 ON fichajes1.IdEmpleado = empleados1.Codigo WHERE  Fecha = '20230213' GROUP BY IdEmpleado HAVING COUNT(*) = 1 ORDER BY IdEmpleado ASC"

# Realiza la consulta
micursor = mydb.cursor()
micursor.execute(consulta)

# Si el archivo ya existe, se carga, de lo contrario se crea uno nuevo

if os.path.exists("C:\\AOBREGON\\"+nombre_archivo):
    archiExiste(micursor)                                              #El archivo ya existe
    print("B")
else:
    archiNoExiste(micursor)
    print("a")                                                         #El archivo no Existe


print(f"Se han guardado las incidencias en el archivo {nombre_archivo}")
mydb.close()
