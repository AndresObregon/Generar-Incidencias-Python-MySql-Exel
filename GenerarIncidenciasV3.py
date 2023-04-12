import mysql.connector
import datetime
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

# Configura la conexión
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="contraseña",
    database="base de datos"
)

#Saber dia de hoy
fecha_hoy = date.today().strftime('%Y%m%d')

# Obtiene el número de la semana actual
semana_actual = datetime.date.today().isocalendar()[1]

# Nombre del archivo de Excel
nombre_archivo = f"ErroresFichajeSemana{semana_actual}.xlsx"

if os.path.exists("C:\\AOBREGON\\"+nombre_archivo): #El archivo ya existe, Lo abrimos y selecionamos la hoja y asignamos la fila aparir de la ultima escrita
    libro = openpyxl.load_workbook("C:\\AOBREGON\\"+nombre_archivo)
    hoja_excel = libro.active
    fila = 1 + hoja_excel.max_row

else: #El archivo no Existe, Creamos el exel, selecionamos la hoja y asiganmos los encabezasos 
    libro = openpyxl.Workbook()
    hoja_excel = libro.active
     # Crea encabezados para las columnas
    hoja_excel.cell(row=1, column=1, value="Fecha").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=2, value="Hora Fichaje").alignment = Alignment(horizontal="center")
    hoja_excel.cell(row=1, column=3, value="Alias").alignment = Alignment(horizontal="center")
    fila = 2

# Consulta para obtener las incidencias
consulta = "SELECT Fecha, MAX(Hora) as HoraFichaje, Alias FROM newschema.fichajes1 JOIN newschema.empleados1 ON fichajes1.IdEmpleado = empleados1.Codigo WHERE  Fecha = '20230213' GROUP BY IdEmpleado HAVING COUNT(*) = 1 ORDER BY IdEmpleado ASC"

# Realiza la consulta
mycursor = mydb.cursor()
mycursor.execute(consulta)

for x in mycursor: # Ajusta el ancho de las columnas
    if len(x[0]) == 8:
        fecha = datetime.datetime.strptime(x[0], '%Y%m%d').strftime('%Y/%m/%d')
        hora_Fichaje = datetime.datetime.strptime(x[1], '%H%M%S').strftime('%H:%M:%S')
        hoja_excel.cell(row=fila, column=1, value=fecha)
        hoja_excel.cell(row=fila, column=2, value=hora_Fichaje)
        hoja_excel.cell(row=fila, column=3, value=x[2])
        fila += 1
    else:
        print(f"Error: Fecha inválida ({x[0]})")
for col in hoja_excel.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    hoja_excel.column_dimensions[column].width = adjusted_width

libro.save("C:\\AOBREGON\\" + nombre_archivo)# Guarda el archivo de Excel


print(f"Se han guardado las incidencias en el archivo {nombre_archivo}")
mydb.close() #Cerrar conexion sql