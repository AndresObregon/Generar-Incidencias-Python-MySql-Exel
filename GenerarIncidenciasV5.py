import mysql.connector
import datetime
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

def ajustarExel(hoja_excel): #Con esto le damos formato a las celdas para que quedebe bien
    for col in hoja_excel.columns: 
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja_excel.column_dimensions[column].width = adjusted_width

# Configura la conexión
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="contraseña",
    database="base de datos"
)

# Consulta para obtener las incidencias
with open('consulta.txt', 'r') as file:
    lineas = file.readlines()
    consulta1 = lineas[0].strip()
    consulta2 = lineas[1].strip()

# Realiza la consulta
Cursor1 = mydb.cursor()
Cursor1.execute(consulta1)
results = Cursor1.fetchall()

Cursor2 = mydb.cursor()
Cursor2.execute(consulta2)
results2 = Cursor2.fetchall()

if len(results) == 0 and len(results2) == 0:
    print("DIA SIN INCIDENCIAS")
else:  
    # Obtiene el número de la semana actual
    semana_actual = datetime.date.today().isocalendar()[1]

    # Nombre del archivo de Excel
    nombre_archivo = f"ErroresFichajeSemana{semana_actual}.xlsx"

    if os.path.exists(nombre_archivo): #El archivo ya existe, Lo abrimos y selecionamos la hoja y asignamos la fila aparir de la ultima escrita
        print("Abrimos Exel")
        libro = openpyxl.load_workbook(nombre_archivo)
        hoja_excel = libro.active
        fila = 1 + hoja_excel.max_row

    else: #El archivo no Existe, Creamos el exel, selecionamos la hoja y asiganmos los encabezasos 
        print("Creamos Archivo Exel")
        libro = openpyxl.Workbook()
        hoja_excel = libro.active
         # Crea encabezados para las columnas
        hoja_excel.cell(row=1, column=1, value="Fecha").alignment = Alignment(horizontal="center")
        hoja_excel.cell(row=1, column=2, value="Hora Fichaje").alignment = Alignment(horizontal="center")
        hoja_excel.cell(row=1, column=3, value="Codigo").alignment = Alignment(horizontal="center")
        hoja_excel.cell(row=1, column=4, value="Alias").alignment = Alignment(horizontal="center")
        hoja_excel.cell(row=1, column=5, value="Tipo Incidencia").alignment = Alignment(horizontal="center")
        fila = 2

    for x in results: # Escribimos los datos del query en el exel
        if len(x[0]) == 8:
            fecha = datetime.datetime.strptime(x[0], '%Y%m%d').strftime('%Y/%m/%d')
            hora_Fichaje = datetime.datetime.strptime(x[1], '%H%M%S').strftime('%H:%M:%S')
            hoja_excel.cell(row=fila, column=1, value=fecha)
            hoja_excel.cell(row=fila, column=2, value=hora_Fichaje)
            hoja_excel.cell(row=fila, column=3, value=x[2])
            hoja_excel.cell(row=fila, column=4, value=x[3])
            hoja_excel.cell(row=fila, column=5, value="Falta entrada/salida")
            fila += 1
        else:
            print(f"Error: Fecha inválida ({x[0]})")
    
    if len(results2) != 0: #Comprobamos si hay usuaraios que no ficharon el dia 
        fecha_hoy = date.today().strftime('%Y/%m/%d')

        for x in results2: #Escribimos las personas que no ficharon este dia
            hoja_excel.cell(row=fila, column=1, value=fecha_hoy)
            hoja_excel.cell(row=fila, column=2, value="NONE")
            hoja_excel.cell(row=fila, column=3, value=x[0])
            hoja_excel.cell(row=fila, column=4, value=x[1])
            hoja_excel.cell(row=fila, column=5, value="No Ficho el dia")
            fila += 1
    ajustarExel(hoja_excel)
    libro.save(nombre_archivo)# Guarda el archivo de Excel
    print(f"Se han guardado las incidencias en el archivo {nombre_archivo}")
mydb.close() #Cerrar conexion sql